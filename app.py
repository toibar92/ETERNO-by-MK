from flask import Flask, render_template, request, redirect, url_for, session, flash, send_file
import os
import psycopg2
from psycopg2.extras import RealDictCursor
from datetime import datetime, timedelta, date
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
import io
import secrets

app = Flask(__name__)

app.secret_key = os.environ.get('SECRET_KEY', secrets.token_hex(32))
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(minutes=30)

if os.environ.get('RENDER'):
    app.config['SESSION_COOKIE_SECURE'] = True

DATABASE_URL = os.environ.get('DATABASE_URL')

# CAMBIO 1: Mediano → Pequeño
PRODUCTOS_CONFIG = {
    'Grande': {
        'medidas': '94 x 152 cms',
        'precio': 15750,
        'costos': {'enmarcado': 1990, 'impresion': 887, 'drytac': 576, 'acrilico': 740, 'montaje': 300, 'papel': 340}
    },
    'Pequeño': {
        'medidas': '75 x 122 cms',
        'precio': 13125,
        'costos': {'enmarcado': 1500, 'impresion': 595, 'drytac': 576, 'acrilico': 470, 'montaje': 300, 'papel': 320}
    }
}

COSTOS_VISA = {2: 0.05, 3: 0.0575, 6: 0.07, 10: 0.07, 12: 0.08}
METODOS_PAGO = ['Efectivo', 'Transferencia', 'Tarjeta débito', 'Pendiente']

def get_client_ip():
    if request.headers.get('X-Forwarded-For'):
        return request.headers.get('X-Forwarded-For').split(',')[0].strip()
    return request.remote_addr or '0.0.0.0'

def log_security_event(event_type, user_id=None, details=""):
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute('INSERT INTO security_log (event_type, user_id, ip_address, details, timestamp) VALUES (%s, %s, %s, %s, NOW())',
                   (event_type, user_id, get_client_ip(), details))
        conn.commit()
        cur.close()
        conn.close()
    except:
        pass

def check_login_attempts(username, ip_address):
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) as attempts FROM login_attempts WHERE username = %s AND success = FALSE AND attempt_time > NOW() - INTERVAL '15 minutes'", (username,))
        result = cur.fetchone()
        cur.close()
        conn.close()
        if result and result['attempts'] >= 5:
            return False, "Demasiados intentos fallidos. Intenta en 15 minutos."
        return True, ""
    except:
        return True, ""

def log_login_attempt(username, ip_address, success):
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute('INSERT INTO login_attempts (username, ip_address, success, attempt_time) VALUES (%s, %s, %s, NOW())', (username, ip_address, success))
        conn.commit()
        cur.close()
        conn.close()
    except:
        pass

def get_db_connection():
    conn = psycopg2.connect(DATABASE_URL, cursor_factory=RealDictCursor)
    return conn

def init_db():
    conn = get_db_connection()
    cur = conn.cursor()
    
    cur.execute('''CREATE TABLE IF NOT EXISTS usuarios (
        id SERIAL PRIMARY KEY, username VARCHAR(50) UNIQUE NOT NULL, password VARCHAR(255) NOT NULL,
        nombre VARCHAR(100) NOT NULL, rol VARCHAR(20) DEFAULT 'usuario', activo BOOLEAN DEFAULT TRUE,
        fecha_creacion TIMESTAMP DEFAULT CURRENT_TIMESTAMP)''')
    
    cur.execute('''CREATE TABLE IF NOT EXISTS pedidos (
        id SERIAL PRIMARY KEY, fecha DATE NOT NULL, cliente VARCHAR(100) NOT NULL, producto VARCHAR(50) NOT NULL,
        cantidad INTEGER DEFAULT 1, precio_unitario DECIMAL(10,2) NOT NULL, descuento DECIMAL(5,4) DEFAULT 0,
        anticipo DECIMAL(10,2) DEFAULT 0, metodo_pago_anticipo VARCHAR(50), cuotas_visa_anticipo INTEGER DEFAULT 0,
        fecha_sesion DATE, metodo_pago_saldo VARCHAR(50), cuotas_visa_saldo INTEGER DEFAULT 0,
        saldo_pagado BOOLEAN DEFAULT FALSE,
        usuario_id INTEGER REFERENCES usuarios(id), fecha_creacion TIMESTAMP DEFAULT CURRENT_TIMESTAMP)''')
    
    cur.execute('''CREATE TABLE IF NOT EXISTS login_attempts (
        id SERIAL PRIMARY KEY, username VARCHAR(50) NOT NULL, ip_address VARCHAR(45) NOT NULL,
        success BOOLEAN NOT NULL, attempt_time TIMESTAMP DEFAULT CURRENT_TIMESTAMP)''')
    
    cur.execute('''CREATE TABLE IF NOT EXISTS security_log (
        id SERIAL PRIMARY KEY, event_type VARCHAR(50) NOT NULL, user_id INTEGER REFERENCES usuarios(id),
        ip_address VARCHAR(45), details TEXT, timestamp TIMESTAMP DEFAULT CURRENT_TIMESTAMP)''')
    
    try:
        cur.execute('ALTER TABLE pedidos ADD COLUMN IF NOT EXISTS anticipo DECIMAL(10,2) DEFAULT 0')
        cur.execute('ALTER TABLE pedidos ADD COLUMN IF NOT EXISTS metodo_pago_anticipo VARCHAR(50)')
        cur.execute('ALTER TABLE pedidos ADD COLUMN IF NOT EXISTS cuotas_visa_anticipo INTEGER DEFAULT 0')
        cur.execute('ALTER TABLE pedidos ADD COLUMN IF NOT EXISTS fecha_sesion DATE')
        cur.execute('ALTER TABLE pedidos ADD COLUMN IF NOT EXISTS metodo_pago_saldo VARCHAR(50)')
        cur.execute('ALTER TABLE pedidos ADD COLUMN IF NOT EXISTS cuotas_visa_saldo INTEGER DEFAULT 0')
        cur.execute('ALTER TABLE pedidos ADD COLUMN IF NOT EXISTS saldo_pagado BOOLEAN DEFAULT FALSE')
    except:
        pass
    
    cur.execute("SELECT * FROM usuarios WHERE username = 'admin'")
    if not cur.fetchone():
        hashed_password = generate_password_hash('Eterno2026!')
        cur.execute('INSERT INTO usuarios (username, password, nombre, rol) VALUES (%s, %s, %s, %s)',
                   ('admin', hashed_password, 'Administrador', 'admin'))
    
    conn.commit()
    cur.close()
    conn.close()

try:
    init_db()
except Exception as e:
    print(f"Error inicializando DB: {e}")

# CAMBIO 2: Validación de fechas para saldos pendientes
def calcular_totales(producto, cantidad, descuento=0, anticipo=0, cuotas_visa_anticipo=0, cuotas_visa_saldo=0, fecha_sesion=None, saldo_pagado=False):
    config = PRODUCTOS_CONFIG[producto]
    precio_unitario = config['precio']
    costos_detalle = {k: v * cantidad for k, v in config['costos'].items()}
    costo_produccion = sum(costos_detalle.values())
    subtotal = precio_unitario * cantidad
    total_venta = subtotal * (1 - descuento)
    saldo_restante = total_venta - anticipo
    costo_visa_anticipo = anticipo * COSTOS_VISA.get(cuotas_visa_anticipo, 0) if cuotas_visa_anticipo > 0 else 0
    costo_visa_saldo = saldo_restante * COSTOS_VISA.get(cuotas_visa_saldo, 0) if cuotas_visa_saldo > 0 else 0
    costo_visa_total = costo_visa_anticipo + costo_visa_saldo
    costo_total = costo_produccion + costo_visa_total
    utilidad = total_venta - costo_total
    porcentaje_utilidad = (utilidad / total_venta * 100) if total_venta > 0 else 0
    disponible_anticipo = anticipo - costo_produccion - costo_visa_anticipo
    
    # NUEVA LÓGICA: Si la fecha de sesión ya pasó O está marcado como pagado, NO es saldo pendiente
    es_saldo_pendiente = False
    if saldo_restante > 0 and not saldo_pagado:
        if fecha_sesion:
            if isinstance(fecha_sesion, str):
                fecha_sesion = datetime.strptime(fecha_sesion, '%Y-%m-%d').date()
            es_saldo_pendiente = fecha_sesion > date.today()
        else:
            es_saldo_pendiente = True
    
    return {
        'precio_unitario': precio_unitario, 'subtotal': subtotal, 'total_venta': total_venta,
        'anticipo': anticipo, 'saldo_restante': saldo_restante, 'costos_detalle': costos_detalle,
        'costo_produccion': costo_produccion, 'costo_visa_anticipo': costo_visa_anticipo,
        'costo_visa_saldo': costo_visa_saldo, 'costo_visa_total': costo_visa_total,
        'costo_total': costo_total, 'utilidad': utilidad, 'porcentaje_utilidad': porcentaje_utilidad,
        'disponible_anticipo': disponible_anticipo, 'es_saldo_pendiente': es_saldo_pendiente
    }

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Debes iniciar sesión', 'warning')
            return redirect(url_for('login'))
        if 'last_activity' in session:
            try:
                last_activity = datetime.fromisoformat(session['last_activity'])
                if datetime.now() - last_activity > timedelta(minutes=30):
                    session.clear()
                    flash('Tu sesión ha expirado', 'warning')
                    return redirect(url_for('login'))
            except:
                session.clear()
                return redirect(url_for('login'))
        session['last_activity'] = datetime.now().isoformat()
        session.modified = True
        return f(*args, **kwargs)
    return decorated_function

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Debes iniciar sesión', 'warning')
            return redirect(url_for('login'))
        if session.get('rol') != 'admin':
            flash('No tienes permisos', 'error')
            return redirect(url_for('dashboard'))
        if 'last_activity' in session:
            try:
                last_activity = datetime.fromisoformat(session['last_activity'])
                if datetime.now() - last_activity > timedelta(minutes=30):
                    session.clear()
                    return redirect(url_for('login'))
            except:
                session.clear()
                return redirect(url_for('login'))
        session['last_activity'] = datetime.now().isoformat()
        session.modified = True
        return f(*args, **kwargs)
    return decorated_function
@app.before_request
def require_login():
  if request.endpoint in ['login', 'static', 'index', None]:
        return None
    if request.path == '/login':
        return None
    if request.path.startswith('/static'):
        return None
    if 'user_id' not in session:
        return redirect(url_for('login'))
@app.route('/')
def index():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        if not username or not password:
            flash('Usuario y contraseña son requeridos', 'error')
            return render_template('login.html')
        ip_address = get_client_ip()
        can_attempt, message = check_login_attempts(username, ip_address)
        if not can_attempt:
            flash(message, 'error')
            log_security_event('login_rate_limit_exceeded', None, f"Usuario: {username}")
            return render_template('login.html')
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute('SELECT * FROM usuarios WHERE username = %s AND activo = TRUE', (username,))
        user = cur.fetchone()
        cur.close()
        conn.close()
        if user and check_password_hash(user['password'], password):
            session.clear()
            session['user_id'] = user['id']
            session['username'] = user['username']
            session['nombre'] = user['nombre']
            session['rol'] = user['rol']
            session['last_activity'] = datetime.now().isoformat()
            session.permanent = True
            log_login_attempt(username, ip_address, True)
            log_security_event('login_success', user['id'])
            flash(f'Bienvenido, {user["nombre"]}!', 'success')
            return redirect(url_for('dashboard'))
        else:
            log_login_attempt(username, ip_address, False)
            log_security_event('login_failed', None, f"Usuario: {username}")
            flash('Usuario o contraseña incorrectos', 'error')
    return render_template('login.html')

@app.route('/logout')
def logout():
    user_id = session.get('user_id')
    log_security_event('logout', user_id)
    session.clear()
    flash('Sesión cerrada exitosamente', 'info')
    return redirect(url_for('login'))

@app.route('/dashboard')
@login_required
def dashboard():
    fecha_inicio = request.args.get('fecha_inicio', '')
    fecha_fin = request.args.get('fecha_fin', '')
    conn = get_db_connection()
    cur = conn.cursor()
    query = 'SELECT * FROM pedidos'
    params = []
    if fecha_inicio and fecha_fin:
        query += ' WHERE fecha BETWEEN %s AND %s'
        params = [fecha_inicio, fecha_fin]
    elif fecha_inicio:
        query += ' WHERE fecha >= %s'
        params = [fecha_inicio]
    elif fecha_fin:
        query += ' WHERE fecha <= %s'
        params = [fecha_fin]
    query += ' ORDER BY fecha DESC, id DESC'
    cur.execute(query, params)
    pedidos = cur.fetchall()
    cur.close()
    conn.close()
    total_anticipos = total_saldos_pendientes = total_costos = total_utilidad = 0
    pedidos_grande = pedidos_pequeno = anticipos_grande = anticipos_pequeno = 0
    saldos_grande = saldos_pequeno = costos_grande = costos_pequeno = 0
    pedidos_procesados = []
    for pedido in pedidos:
        try:
            anticipo = float(pedido['anticipo']) if pedido.get('anticipo') else 0
            totales = calcular_totales(pedido['producto'], pedido['cantidad'],
                                      float(pedido['descuento']) if pedido['descuento'] else 0,
                                      anticipo, pedido.get('cuotas_visa_anticipo') or 0,
                                      pedido.get('cuotas_visa_saldo') or 0,
                                      pedido.get('fecha_sesion'), pedido.get('saldo_pagado', False))
            pedidos_procesados.append({'id': pedido['id'], 'fecha': pedido['fecha'], 'cliente': pedido['cliente'],
                                      'producto': pedido['producto'], 'cantidad': pedido['cantidad'],
                                      'fecha_sesion': pedido.get('fecha_sesion'),
                                      'metodo_pago_anticipo': pedido.get('metodo_pago_anticipo', ''),
                                      'metodo_pago_saldo': pedido.get('metodo_pago_saldo', ''),
                                      'saldo_pagado': pedido.get('saldo_pagado', False), **totales})
            total_anticipos += anticipo
            if totales['es_saldo_pendiente']:
                total_saldos_pendientes += totales['saldo_restante']
            total_costos += totales['costo_total']
            total_utilidad += totales['utilidad']
            if pedido['producto'] == 'Grande':
                pedidos_grande += pedido['cantidad']
                anticipos_grande += anticipo
                if totales['es_saldo_pendiente']:
                    saldos_grande += totales['saldo_restante']
                costos_grande += totales['costo_total']
            else:
                pedidos_pequeno += pedido['cantidad']
                anticipos_pequeno += anticipo
                if totales['es_saldo_pendiente']:
                    saldos_pequeno += totales['saldo_restante']
                costos_pequeno += totales['costo_total']
        except:
            continue
    total_ventas_proyectadas = total_anticipos + total_saldos_pendientes
    margen_promedio = (total_utilidad / total_ventas_proyectadas * 100) if total_ventas_proyectadas > 0 else 0
    estadisticas = {
        'total_anticipos': total_anticipos, 'total_saldos_pendientes': total_saldos_pendientes,
        'total_ventas_proyectadas': total_ventas_proyectadas, 'total_costos': total_costos,
        'total_utilidad': total_utilidad, 'margen_promedio': margen_promedio, 'total_pedidos': len(pedidos_procesados),
        'pedidos_grande': pedidos_grande, 'pedidos_pequeno': pedidos_pequeno,
        'anticipos_grande': anticipos_grande, 'anticipos_pequeno': anticipos_pequeno,
        'saldos_grande': saldos_grande, 'saldos_pequeno': saldos_pequeno,
        'costos_grande': costos_grande, 'costos_pequeno': costos_pequeno
    }
    return render_template('dashboard.html', pedidos=pedidos_procesados, estadisticas=estadisticas,
                         fecha_inicio=fecha_inicio, fecha_fin=fecha_fin)

# NUEVA RUTA: Módulo de Saldos Pendientes
@app.route('/saldos-pendientes')
@login_required
def saldos_pendientes():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute('''SELECT * FROM pedidos 
                   WHERE ((fecha_sesion IS NULL OR fecha_sesion > CURRENT_DATE) AND saldo_pagado = FALSE)
                   ORDER BY fecha_sesion ASC NULLS LAST, fecha DESC''')
    pedidos = cur.fetchall()
    cur.close()
    conn.close()
    
    pedidos_pendientes = []
    for pedido in pedidos:
        try:
            anticipo = float(pedido['anticipo']) if pedido.get('anticipo') else 0
            totales = calcular_totales(pedido['producto'], pedido['cantidad'],
                                      float(pedido['descuento']) if pedido['descuento'] else 0,
                                      anticipo, pedido.get('cuotas_visa_anticipo') or 0,
                                      pedido.get('cuotas_visa_saldo') or 0,
                                      pedido.get('fecha_sesion'), False)
            if totales['saldo_restante'] > 0:
                pedidos_pendientes.append({
                    'id': pedido['id'],
                    'fecha': pedido['fecha'],
                    'cliente': pedido['cliente'],
                    'producto': pedido['producto'],
                    'cantidad': pedido['cantidad'],
                    'fecha_sesion': pedido.get('fecha_sesion'),
                    'metodo_pago_saldo': pedido.get('metodo_pago_saldo', 'Pendiente'),
                    'cuotas_visa_saldo': pedido.get('cuotas_visa_saldo', 0),
                    **totales
                })
        except:
            continue
    
    return render_template('saldos_pendientes.html', pedidos=pedidos_pendientes, metodos_pago=METODOS_PAGO, costos_visa=COSTOS_VISA)

# NUEVA RUTA: Marcar saldo como pagado
@app.route('/marcar-saldo-pagado/<int:pedido_id>', methods=['POST'])
@login_required
def marcar_saldo_pagado(pedido_id):
    metodo_pago = request.form.get('metodo_pago_saldo', '')
    cuotas_visa = int(request.form.get('cuotas_visa_saldo', 0))
    
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute('''UPDATE pedidos SET saldo_pagado = TRUE, metodo_pago_saldo = %s, cuotas_visa_saldo = %s 
                   WHERE id = %s''', (metodo_pago, cuotas_visa, pedido_id))
    conn.commit()
    cur.close()
    conn.close()
    
    flash('Saldo marcado como pagado exitosamente', 'success')
    return redirect(url_for('saldos_pendientes'))

@app.route('/exportar-excel')
@login_required
def exportar_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    fecha_inicio = request.args.get('fecha_inicio', '')
    fecha_fin = request.args.get('fecha_fin', '')
    conn = get_db_connection()
    cur = conn.cursor()
    query = 'SELECT * FROM pedidos'
    params = []
    if fecha_inicio and fecha_fin:
        query += ' WHERE fecha BETWEEN %s AND %s'
        params = [fecha_inicio, fecha_fin]
    elif fecha_inicio:
        query += ' WHERE fecha >= %s'
        params = [fecha_inicio]
    elif fecha_fin:
        query += ' WHERE fecha <= %s'
        params = [fecha_fin]
    query += ' ORDER BY fecha DESC'
    cur.execute(query, params)
    pedidos = cur.fetchall()
    cur.close()
    conn.close()
    wb = Workbook()
    ws = wb.active
    ws.title = "Pedidos"
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    cell_alignment = Alignment(horizontal="center", vertical="center")
    money_alignment = Alignment(horizontal="right", vertical="center")
    thin_border = Border(left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'),
                        top=Side(style='thin', color='CCCCCC'), bottom=Side(style='thin', color='CCCCCC'))
    green_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    orange_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
    red_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
    ws.merge_cells('A1:J1')
    ws['A1'] = 'ETERNO by MK - Reporte de Pedidos'
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal="center")
    ws.merge_cells('A2:J2')
    fecha_reporte = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    if fecha_inicio and fecha_fin:
        fecha_reporte += f" | Período: {fecha_inicio} a {fecha_fin}"
    ws['A2'] = fecha_reporte
    ws['A2'].alignment = Alignment(horizontal="center")
    ws['A2'].font = Font(size=10, color="666666")
    headers = ['Fecha', 'Cliente', 'Producto', 'Cant.', 'Total Venta', 'Anticipo', 'Saldo', 'Fecha Sesión', 'Costo', 'Utilidad']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    row = 5
    total_venta = total_anticipo = total_saldo = total_costo = total_utilidad = 0
    for pedido in pedidos:
        try:
            anticipo = float(pedido['anticipo']) if pedido.get('anticipo') else 0
            totales = calcular_totales(pedido['producto'], pedido['cantidad'],
                                      float(pedido['descuento']) if pedido['descuento'] else 0, anticipo,
                                      pedido.get('cuotas_visa_anticipo') or 0, pedido.get('cuotas_visa_saldo') or 0,
                                      pedido.get('fecha_sesion'), pedido.get('saldo_pagado', False))
            ws.cell(row=row, column=1, value=pedido['fecha'].strftime('%d/%m/%Y') if pedido['fecha'] else '').alignment = cell_alignment
            ws.cell(row=row, column=2, value=pedido['cliente'])
            ws.cell(row=row, column=3, value=pedido['producto']).alignment = cell_alignment
            ws.cell(row=row, column=4, value=pedido['cantidad']).alignment = cell_alignment
            cell_venta = ws.cell(row=row, column=5, value=totales['total_venta'])
            cell_venta.number_format = '"Q"#,##0.00'
            cell_venta.alignment = money_alignment
            cell_anticipo = ws.cell(row=row, column=6, value=anticipo)
            cell_anticipo.number_format = '"Q"#,##0.00'
            cell_anticipo.alignment = money_alignment
            cell_anticipo.fill = green_fill
            cell_saldo = ws.cell(row=row, column=7, value=totales['saldo_restante'])
            cell_saldo.number_format = '"Q"#,##0.00'
            cell_saldo.alignment = money_alignment
            cell_saldo.fill = orange_fill
            ws.cell(row=row, column=8, value=pedido['fecha_sesion'].strftime('%d/%m/%Y') if pedido.get('fecha_sesion') else '-').alignment = cell_alignment
            cell_costo = ws.cell(row=row, column=9, value=totales['costo_total'])
            cell_costo.number_format = '"Q"#,##0.00'
            cell_costo.alignment = money_alignment
            cell_costo.fill = red_fill
            cell_utilidad = ws.cell(row=row, column=10, value=totales['utilidad'])
            cell_utilidad.number_format = '"Q"#,##0.00'
            cell_utilidad.alignment = money_alignment
            for col in range(1, 11):
                ws.cell(row=row, column=col).border = thin_border
            total_venta += totales['total_venta']
            total_anticipo += anticipo
            total_saldo += totales['saldo_restante']
            total_costo += totales['costo_total']
            total_utilidad += totales['utilidad']
            row += 1
        except:
            continue
    row += 1
    ws.cell(row=row, column=4, value='TOTALES:').font = Font(bold=True)
    for col, val in [(5, total_venta), (6, total_anticipo), (7, total_saldo), (9, total_costo), (10, total_utilidad)]:
        cell = ws.cell(row=row, column=col, value=val)
        cell.number_format = '"Q"#,##0.00'
        cell.font = Font(bold=True)
        cell.alignment = money_alignment
    column_widths = [12, 25, 12, 8, 15, 15, 15, 14, 15, 15]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"pedidos_eterno_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True, download_name=filename)

@app.route('/nuevo-pedido', methods=['GET', 'POST'])
@login_required
def nuevo_pedido():
    if request.method == 'POST':
        fecha = request.form['fecha']
        cliente = request.form['cliente']
        producto = request.form['producto']
        cantidad = int(request.form['cantidad'])
        descuento = float(request.form.get('descuento', 0)) / 100
        anticipo = float(request.form.get('anticipo', 0))
        metodo_pago_anticipo = request.form.get('metodo_pago_anticipo', '')
        cuotas_visa_anticipo = int(request.form.get('cuotas_visa_anticipo', 0))
        fecha_sesion = request.form.get('fecha_sesion') or None
        metodo_pago_saldo = request.form.get('metodo_pago_saldo', '')
        cuotas_visa_saldo = int(request.form.get('cuotas_visa_saldo', 0))
        precio_unitario = PRODUCTOS_CONFIG[producto]['precio']
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute('''INSERT INTO pedidos (fecha, cliente, producto, cantidad, precio_unitario, descuento, anticipo, 
                      metodo_pago_anticipo, cuotas_visa_anticipo, fecha_sesion, metodo_pago_saldo, cuotas_visa_saldo, usuario_id)
                      VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)''',
                   (fecha, cliente, producto, cantidad, precio_unitario, descuento, anticipo, metodo_pago_anticipo,
                    cuotas_visa_anticipo, fecha_sesion, metodo_pago_saldo, cuotas_visa_saldo, session['user_id']))
        conn.commit()
        cur.close()
        conn.close()
        flash('Pedido registrado exitosamente', 'success')
        return redirect(url_for('dashboard'))
    return render_template('nuevo_pedido.html', productos=PRODUCTOS_CONFIG, costos_visa=COSTOS_VISA, metodos_pago=METODOS_PAGO)

@app.route('/editar-pedido/<int:pedido_id>', methods=['GET', 'POST'])
@login_required
def editar_pedido(pedido_id):
    conn = get_db_connection()
    cur = conn.cursor()
    if request.method == 'POST':
        fecha = request.form['fecha']
        cliente = request.form['cliente']
        producto = request.form['producto']
        cantidad = int(request.form['cantidad'])
        descuento = float(request.form.get('descuento', 0)) / 100
        anticipo = float(request.form.get('anticipo', 0))
        metodo_pago_anticipo = request.form.get('metodo_pago_anticipo', '')
        cuotas_visa_anticipo = int(request.form.get('cuotas_visa_anticipo', 0))
        fecha_sesion = request.form.get('fecha_sesion') or None
        metodo_pago_saldo = request.form.get('metodo_pago_saldo', '')
        cuotas_visa_saldo = int(request.form.get('cuotas_visa_saldo', 0))
        precio_unitario = PRODUCTOS_CONFIG[producto]['precio']
        cur.execute('''UPDATE pedidos SET fecha=%s, cliente=%s, producto=%s, cantidad=%s, precio_unitario=%s, descuento=%s, 
                      anticipo=%s, metodo_pago_anticipo=%s, cuotas_visa_anticipo=%s, fecha_sesion=%s, metodo_pago_saldo=%s, 
                      cuotas_visa_saldo=%s WHERE id=%s''',
                   (fecha, cliente, producto, cantidad, precio_unitario, descuento, anticipo, metodo_pago_anticipo,
                    cuotas_visa_anticipo, fecha_sesion, metodo_pago_saldo, cuotas_visa_saldo, pedido_id))
        conn.commit()
        cur.close()
        conn.close()
        flash('Pedido actualizado', 'success')
        return redirect(url_for('dashboard'))
    cur.execute('SELECT * FROM pedidos WHERE id = %s', (pedido_id,))
    pedido = cur.fetchone()
    cur.close()
    conn.close()
    return render_template('editar_pedido.html', pedido=pedido, productos=PRODUCTOS_CONFIG, costos_visa=COSTOS_VISA, metodos_pago=METODOS_PAGO)

@app.route('/eliminar-pedido/<int:pedido_id>')
@login_required
def eliminar_pedido(pedido_id):
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute('DELETE FROM pedidos WHERE id = %s', (pedido_id,))
    conn.commit()
    cur.close()
    conn.close()
    flash('Pedido eliminado', 'success')
    return redirect(url_for('dashboard'))

@app.route('/usuarios')
@admin_required
def usuarios():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute('SELECT * FROM usuarios ORDER BY id')
    users = cur.fetchall()
    cur.close()
    conn.close()
    return render_template('usuarios.html', usuarios=users)

@app.route('/usuarios/nuevo', methods=['GET', 'POST'])
@admin_required
def nuevo_usuario():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        nombre = request.form['nombre']
        rol = request.form.get('rol', 'usuario')
        hashed_password = generate_password_hash(password)
        conn = get_db_connection()
        cur = conn.cursor()
        try:
            cur.execute('INSERT INTO usuarios (username, password, nombre, rol) VALUES (%s, %s, %s, %s)',
                       (username, hashed_password, nombre, rol))
            conn.commit()
            flash('Usuario creado', 'success')
        except:
            flash('El usuario ya existe', 'error')
        cur.close()
        conn.close()
        return redirect(url_for('usuarios'))
    return render_template('nuevo_usuario.html')

@app.route('/usuarios/editar/<int:id>', methods=['GET', 'POST'])
@admin_required
def editar_usuario(id):
    conn = get_db_connection()
    cur = conn.cursor()
    if request.method == 'POST':
        nombre = request.form['nombre']
        rol = request.form.get('rol', 'usuario')
        activo = 'activo' in request.form
        cur.execute('UPDATE usuarios SET nombre=%s, rol=%s, activo=%s WHERE id=%s', (nombre, rol, activo, id))
        conn.commit()
        cur.close()
        conn.close()
        flash('Usuario actualizado', 'success')
        return redirect(url_for('usuarios'))
    cur.execute('SELECT * FROM usuarios WHERE id = %s', (id,))
    usuario = cur.fetchone()
    cur.close()
    conn.close()
    return render_template('editar_usuario.html', usuario=usuario)

@app.route('/usuarios/eliminar/<int:id>')
@admin_required
def eliminar_usuario(id):
    if id == session['user_id']:
        flash('No puedes eliminarte', 'error')
        return redirect(url_for('usuarios'))
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute('DELETE FROM usuarios WHERE id = %s', (id,))
    conn.commit()
    cur.close()
    conn.close()
    flash('Usuario eliminado', 'success')
    return redirect(url_for('usuarios'))

@app.route('/cambiar-contrasena', methods=['GET', 'POST'])
@login_required
def cambiar_contrasena():
    if request.method == 'POST':
        actual = request.form['password_actual']
        nueva = request.form['password_nueva']
        confirmar = request.form['password_confirmar']
        if nueva != confirmar:
            flash('Las contraseñas no coinciden', 'error')
            return redirect(url_for('cambiar_contrasena'))
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute('SELECT password FROM usuarios WHERE id = %s', (session['user_id'],))
        user = cur.fetchone()
        if not check_password_hash(user['password'], actual):
            flash('Contraseña actual incorrecta', 'error')
            cur.close()
            conn.close()
            return redirect(url_for('cambiar_contrasena'))
        hashed = generate_password_hash(nueva)
        cur.execute('UPDATE usuarios SET password = %s WHERE id = %s', (hashed, session['user_id']))
        conn.commit()
        cur.close()
        conn.close()
        flash('Contraseña actualizada', 'success')
        return redirect(url_for('dashboard'))
    return render_template('cambiar_contrasena.html')

if __name__ == '__main__':
    app.run(debug=True)
